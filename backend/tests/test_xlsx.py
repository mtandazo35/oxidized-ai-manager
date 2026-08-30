import io

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook

from app import main as main_module


pytestmark = pytest.mark.anyio


def client(headers=None) -> AsyncClient:
    transport = ASGITransport(app=main_module.app)
    return AsyncClient(transport=transport, base_url="http://test", headers=headers)


def build_xlsx(rows) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def test_convert_requires_login() -> None:
    async with client() as api:
        response = await api.post(
            "/api/devices/convert-xlsx",
            files={"file": ("r.xlsx", build_xlsx([["a", "b"]]))},
        )

    assert response.status_code == 401


async def test_convert_xlsx_to_rows(auth_headers) -> None:
    content = build_xlsx([
        ["nombre", "ip", "puerto", "usuario", "clave", "grupo"],
        ["rb-core-01", "192.0.2.10", 2222, "backup", "Clave", "EmpresaA"],
        [None, None, None],
        ["rb-core-02", "192.0.2.20"],
    ])
    async with client(auth_headers) as api:
        response = await api.post(
            "/api/devices/convert-xlsx",
            files={"file": ("routers.xlsx", content)},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["rows"] == 3
    lines = body["text"].splitlines()
    assert lines[0] == "nombre,ip,puerto,usuario,clave,grupo"
    assert lines[1] == "rb-core-01,192.0.2.10,2222,backup,Clave,EmpresaA"
    assert lines[2] == "rb-core-02,192.0.2.20"


async def test_template_download_is_valid_xlsx(auth_headers) -> None:
    from openpyxl import load_workbook

    async with client(auth_headers) as api:
        response = await api.get("/api/devices/import-template")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxml")
    workbook = load_workbook(io.BytesIO(response.content))
    header = [cell.value for cell in workbook.active[1]]
    assert header == ["nombre", "ip", "puerto", "usuario", "clave", "grupo"]


async def test_template_requires_login() -> None:
    async with client() as api:
        response = await api.get("/api/devices/import-template")

    assert response.status_code == 401


async def test_convert_rejects_non_xlsx(auth_headers) -> None:
    async with client(auth_headers) as api:
        response = await api.post(
            "/api/devices/convert-xlsx",
            files={"file": ("routers.xlsx", b"esto no es un xlsx")},
        )

    assert response.status_code == 422
