import io

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pydantic import ValidationError

from .auth import current_user
from .config import get_settings
from .repository import DeviceRepository, DuplicateDeviceError
from .scheduler import trigger_node_backup
from .schemas import (
    DeviceCreate,
    DeviceImportRequest,
    DeviceImportResult,
    DeviceOut,
    DeviceUpdate,
)


router = APIRouter(
    prefix="/api/devices",
    tags=["devices"],
    dependencies=[Depends(current_user)],
)


def _repository(request: Request) -> DeviceRepository:
    return request.app.state.devices


@router.get("", response_model=list[DeviceOut])
async def list_devices(request: Request) -> list[dict]:
    return await _repository(request).list_devices()


@router.post("", response_model=DeviceOut, status_code=status.HTTP_201_CREATED)
async def create_device(request: Request, payload: DeviceCreate) -> dict:
    try:
        return await _repository(request).create_device(payload.model_dump())
    except DuplicateDeviceError:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A device with this name already exists.",
        )


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@router.get("/import-template")
async def import_template() -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routers"
    sheet.append(["nombre", "ip", "puerto", "usuario", "clave", "grupo"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.append(["# Borre las filas de ejemplo que empiezan con #"])
    sheet.append(["#rb-core-01", "192.0.2.10", 2222, "backup", "ClaveSegura", "EmpresaA"])
    sheet.append(["#rb-sucursal-02", "192.0.2.20", 22, "backup", "ClaveSegura", "EmpresaB"])
    for column, width in zip("ABCDEF", (22, 18, 10, 16, 18, 18)):
        sheet.column_dimensions[column].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": 'attachment; filename="plantilla-routers.xlsx"'
        },
    )


@router.get("/{device_id}", response_model=DeviceOut)
async def get_device(request: Request, device_id: int) -> dict:
    device = await _repository(request).get_device(device_id)
    if device is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found.")
    return device


@router.patch("/{device_id}", response_model=DeviceOut)
async def update_device(request: Request, device_id: int, payload: DeviceUpdate) -> dict:
    data = payload.model_dump(exclude_unset=True)
    try:
        device = await _repository(request).update_device(device_id, data)
    except DuplicateDeviceError:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A device with this name already exists.",
        )
    if device is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found.")
    return device


@router.delete("/{device_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_device(request: Request, device_id: int) -> None:
    deleted = await _repository(request).delete_device(device_id)
    if not deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found.")


HEADER_WORDS = {"name", "nombre", "host", "router"}


def _parse_import_line(line: str) -> dict:
    for delimiter in (";", "\t"):
        line = line.replace(delimiter, ",")
    parts = [part.strip() for part in line.split(",")]
    if len(parts) < 2:
        raise ValueError("Se esperan al menos nombre e IP separados por coma.")
    data = {"name": parts[0], "address": parts[1]}
    if len(parts) > 2 and parts[2]:
        if not parts[2].isdigit():
            raise ValueError(f"Puerto no numérico: {parts[2]!r}.")
        data["port"] = int(parts[2])
    if len(parts) > 3:
        data["username"] = parts[3]
    if len(parts) > 4:
        data["password"] = parts[4]
    if len(parts) > 5:
        data["group_name"] = parts[5]
    return data


@router.post("/import", response_model=DeviceImportResult)
async def import_devices(
    request: Request, payload: DeviceImportRequest
) -> dict:
    created = 0
    duplicates: list[str] = []
    errors: list[dict] = []
    for line_number, raw_line in enumerate(payload.text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        first_field = line.split(",")[0].split(";")[0].strip().lower()
        if line_number == 1 and first_field in HEADER_WORDS:
            continue
        try:
            device = DeviceCreate(**_parse_import_line(line))
        except (ValueError, ValidationError) as error:
            if isinstance(error, ValidationError):
                field = error.errors()[0].get("loc", ["?"])[0]
                message = f"Valor inválido en el campo '{field}'."
            else:
                message = str(error)
            errors.append({"line": line_number, "message": message})
            continue
        try:
            await _repository(request).create_device(device.model_dump())
            created += 1
        except DuplicateDeviceError:
            duplicates.append(device.name)
    return {"created": created, "duplicates": duplicates, "errors": errors}


MAX_XLSX_BYTES = 5 * 1024 * 1024
MAX_XLSX_ROWS = 5000


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace(",", " ")


@router.post("/convert-xlsx")
async def convert_xlsx(file: UploadFile = File(...)) -> dict:
    content = await file.read()
    if len(content) > MAX_XLSX_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_CONTENT_TOO_LARGE,
            detail="El archivo supera 5 MB.",
        )
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="No se pudo leer el archivo; guárdelo como .xlsx (Excel moderno).",
        )
    lines: list[str] = []
    try:
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if len(lines) >= MAX_XLSX_ROWS:
                break
            cells = [_cell_to_text(value) for value in row]
            while cells and cells[-1] == "":
                cells.pop()
            if not cells:
                continue
            lines.append(",".join(cells))
    finally:
        workbook.close()
    return {"text": "\n".join(lines), "rows": len(lines)}


@router.post("/{device_id}/backup", status_code=status.HTTP_202_ACCEPTED)
async def backup_now(request: Request, device_id: int) -> dict:
    device = await _repository(request).get_device(device_id)
    if device is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found.")
    try:
        await trigger_node_backup(get_settings().oxidized_url, device["name"])
    except httpx.HTTPError:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Oxidized no aceptó la solicitud de respaldo.",
        )
    return {"status": "queued", "node": device["name"]}
